import re
from pathlib import Path
from openpyxl import load_workbook

md_path = Path('docs/Deepkamal_Net_Price_List_010426.md')
excel_path = Path('docs/Deepkamal_Net_Price_List_010426.xlsx')


def normalize(text):
    if text is None:
        return ''
    return re.sub(r'\s+', ' ', str(text).strip()).upper()


def parse_md_table(path):
    text = path.read_text(encoding='utf-8', errors='replace')
    lines = [line.rstrip() for line in text.splitlines() if line.strip()]
    header_idx = None
    for i, line in enumerate(lines):
        if line.startswith('|') and i + 1 < len(lines) and lines[i + 1].startswith('|') and '---' in lines[i + 1]:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError('Markdown table header not found')

    header = [cell.strip() for cell in lines[header_idx].strip('|').split('|')]
    rows = []
    for line in lines[header_idx + 2:]:
        if not line.startswith('|'):
            break
        cells = [cell.strip() for cell in line.strip('|').split('|')]
        while len(cells) < len(header):
            cells.append('')
        rows.append(dict(zip(header, cells)))
    return rows


md_rows = parse_md_table(md_path)
if not md_rows:
    raise SystemExit('No rows found in markdown table')

lookup = {}
for row in md_rows:
    key = (
        normalize(row.get('CATEGORY / MACHINE')),
        normalize(row.get('PART NO & DESCRIPTION')),
        normalize(row.get('NET RATE (₹)')),
    )
    if key in lookup:
        # prefer first occurrence if duplicates exist
        continue
    lookup[key] = {
        'POPrice(£)': row.get('POPrice(£)', ''),
        'ImportCost(£)': row.get('ImportCost(£)', ''),
        'Total Cost (£)': row.get('Total Cost (£)', ''),
    }

wb = load_workbook(excel_path)
ws = wb['Net Price List'] if 'Net Price List' in wb.sheetnames else wb.active
header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
header_norm = [normalize(value) for value in header]

needed = ['POPrice(£)', 'ImportCost(£)', 'Total Cost (£)']
missing = [col for col in needed if col not in header]
exp_pos = next((i for i, value in enumerate(header)
               if normalize(value) == 'EXPRICE'), None)
insert_at = exp_pos if exp_pos is not None else len(header)
for col in reversed(missing):
    header.insert(insert_at, col)
    ws.insert_cols(insert_at + 1)
    ws.cell(row=1, column=insert_at + 1, value=col)

# Rebuild header positions after insertions
header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_index = {normalize(name): idx + 1 for idx, name in enumerate(header)}

category_col = col_index.get(normalize('CATEGORY / MACHINE'))
part_col = col_index.get(normalize('PART NO & DESCRIPTION'))
net_col = col_index.get(normalize('NET RATE (₹)'))
po_col = col_index.get(normalize('POPRICE(£)'))
import_col = col_index.get(normalize('IMPORTCOST(£)'))
total_col = col_index.get(normalize('TOTAL COST (£)'))

if not (category_col and part_col and net_col and po_col and import_col and total_col):
    raise SystemExit('Failed to locate required columns in Excel header')

updated = 0
for row in range(2, ws.max_row + 1):
    key = (
        normalize(ws.cell(row=row, column=category_col).value),
        normalize(ws.cell(row=row, column=part_col).value),
        normalize(ws.cell(row=row, column=net_col).value),
    )
    if key not in lookup:
        continue
    source = lookup[key]
    for col_name, col_idx in [('POPrice(£)', po_col), ('ImportCost(£)', import_col), ('Total Cost (£)', total_col)]:
        if not ws.cell(row=row, column=col_idx).value:
            value = source.get(col_name, '')
            if value != '':
                ws.cell(row=row, column=col_idx, value=value)
                updated += 1

wb.save(excel_path)
print(f'Updated {updated} missing cell values from markdown backup.')
