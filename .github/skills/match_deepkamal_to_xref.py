import re
import csv
from collections import defaultdict
from pathlib import Path
from difflib import SequenceMatcher
from openpyxl import load_workbook

# Paths
deepkamal_md = Path('docs/Deepkamal_Net_Price_List_010426.md')
xref_md = Path('docs/Durforge_UK-Market-xRef-Price_List_2026.md')
excel_path = Path('docs/Deepkamal_Net_Price_List_010426.xlsx')

# Parse markdown table into rows of dicts


def parse_md_table(path):
    text = path.read_text(encoding='utf-8', errors='replace')
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    # find first table header line and its separator
    header_idx = None
    for i, line in enumerate(lines):
        if line.startswith('|') and '---' in lines[i+1] if i + 1 < len(lines) else False:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f'No markdown table header found in {path}')
    header = [cell.strip() for cell in lines[header_idx].strip('|').split('|')]
    rows = []
    for line in lines[header_idx + 2:]:
        if not line.startswith('|'):
            break
        cells = [cell.strip() for cell in line.strip('|').split('|')]
        # pad to header length
        while len(cells) < len(header):
            cells.append('')
        rows.append(dict(zip(header, cells)))
    return rows


def normalize(text):
    if text is None:
        return ''
    text = str(text).upper().strip()
    text = re.sub(r'\s+', ' ', text)
    return text


def norm_alnum(text):
    return re.sub(r'[^A-Z0-9]', '', normalize(text))


def extract_tokens(text):
    tokens = re.findall(r'[A-Z0-9]+(?:[\-/][A-Z0-9]+)*', normalize(text))
    return [t for t in tokens if t]


def numeric_tokens(text):
    return re.findall(r'[0-9]{3,}', text)


def score_match(src, tgt):
    if not src or not tgt:
        return 0.0
    return SequenceMatcher(None, src, tgt).ratio()


# parse source and target tables
source_rows = parse_md_table(deepkamal_md)
target_rows = parse_md_table(xref_md)
print('source rows', len(source_rows), 'target rows', len(target_rows))

# Build token and numeric indexes for the target table
sku_index = defaultdict(list)
token_index = defaultdict(set)
number_index = defaultdict(set)
for idx, target in enumerate(target_rows):
    t_sku = normalize(target.get('sku', ''))
    t_desc = normalize(target.get('description', ''))
    t_prod = normalize(target.get('product_type', ''))
    t_combined = ' '.join([t_sku, t_desc, t_prod]).strip()
    target['norm_sku'] = norm_alnum(t_sku)
    target['norm_desc'] = norm_alnum(t_desc)
    target['norm_prod'] = norm_alnum(t_prod)
    target['combined'] = t_combined
    target['tokens'] = set(extract_tokens(t_combined))
    target['numbers'] = set(numeric_tokens(
        target['norm_sku'] + t_desc + t_prod))

    if t_sku:
        sku_index[t_sku].append(idx)
    for token in target['tokens']:
        token_index[token].add(idx)
    for number in target['numbers']:
        number_index[number].add(idx)

# Try matching
matches = []
for i, row in enumerate(source_rows):
    part_desc = normalize(row.get('PART NO & DESCRIPTION', ''))
    if not part_desc:
        continue
    src_norm = norm_alnum(part_desc)
    src_tokens = set(extract_tokens(part_desc))
    src_numbers = set(numeric_tokens(src_norm))

    number_candidates = None
    for number in src_numbers:
        ids = number_index.get(number, set())
        if number_candidates is None:
            number_candidates = ids.copy()
        else:
            number_candidates &= ids

    candidate_ids = set()
    if src_norm in sku_index:
        candidate_ids.update(sku_index[src_norm])
    if number_candidates:
        candidate_ids.update(number_candidates)
    elif src_tokens:
        for token in src_tokens:
            candidate_ids.update(token_index.get(token, set()))

    best = None
    best_score = 0.0

    if candidate_ids:
        for idx in candidate_ids:
            target = target_rows[idx]
            if not (target['norm_sku'] or target['norm_desc'] or target['norm_prod']):
                continue
            if src_norm and (src_norm == target['norm_sku'] or src_norm in target['norm_sku'] or target['norm_sku'] in src_norm or src_norm == target['norm_desc'] or src_norm == target['norm_prod']):
                score = 1.0
            else:
                score = max(score_match(src_norm, target['norm_sku']), score_match(
                    src_norm, target['norm_desc']), score_match(src_norm, target['norm_prod']))
            if score >= 0.6 and score > best_score:
                best_score = score
                best = target

    if best_score < 0.9 and src_numbers:
        search_ids = number_candidates if number_candidates is not None else candidate_ids
        if not search_ids:
            search_ids = range(len(target_rows))
        for idx in search_ids:
            target = target_rows[idx]
            if src_numbers and src_numbers.issubset(target['numbers']):
                score = 0.95
                if score > best_score:
                    best_score = score
                    best = target

    if best_score >= 0.9:
        matches.append((i, row, best_score, best))
        continue

    if best and best_score >= 0.80:
        matches.append((i, row, best_score, best))

# Print top matches for review
print('matches', len(matches))
for i, row, score, best in matches[:40]:
    print(
        f"{i}: score={score:.3f} src=\"{row.get('PART NO & DESCRIPTION')}\" "
        f"target_sku=\"{best.get('sku')}\" "
        f"target_desc=\"{best.get('description')}\" "
        f"price={best.get('price_per_unit')}"
    )

# Save matched prices into Excel using openpyxl so formatting is preserved
if matches:
    wb = load_workbook(excel_path)
    sheet_name = 'Net Price List'
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    header_row = [cell.value for cell in ws[1]]
    header_norm = [normalize(value) for value in header_row]
    part_col = next((idx for idx, value in enumerate(header_norm, start=1)
                     if value == 'PARTNO&DESCRIPTION'), 2)

    exp_cols = [idx for idx, value in enumerate(
        header_norm, start=1) if value == 'EXPRICE']
    if exp_cols:
        exp_col = exp_cols[0]
        # remove duplicate ExpPrice columns after the first
        for dup_col in reversed(exp_cols[1:]):
            ws.delete_cols(dup_col)
    else:
        exp_col = len(header_row) + 1
        ws.cell(row=1, column=exp_col, value='ExpPrice')

    part_index = {}
    for row in range(2, ws.max_row + 1):
        part_value = normalize(ws.cell(row=row, column=part_col).value)
        if part_value:
            part_index[part_value] = row

    for i, row, score, best in matches:
        src_part = normalize(row.get('PART NO & DESCRIPTION', ''))
        excel_row = part_index.get(src_part)
        if excel_row:
            price = best.get('price_per_unit', '')
            try:
                price_value = float(price)
            except (TypeError, ValueError):
                price_value = price
            ws.cell(row=excel_row, column=exp_col, value=price_value)

    wb.save(excel_path)
    print('Excel updated with ExpPrice')
else:
    print('No confident matches found; Excel unchanged.')
